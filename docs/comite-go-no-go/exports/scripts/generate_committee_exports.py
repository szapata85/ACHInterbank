from __future__ import annotations

import re
import subprocess
import sys
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Iterable, List, Sequence, Tuple


PROJECT = "ACH Interbank"
PACKAGE_VERSION = "1.0 preliminar"
GLOBAL_STATE = "UAT controlado / NO-GO productivo"
RECOMMENDATION = "Continuar UAT controlado con observaciones"
SCORECARD = "67.8 / 100"
ENVIRONMENT = "Local Docker / UAT controlado"

SOURCE_DOCS = [
    "README.md",
    "00_RESUMEN_EJECUTIVO_COMITE.md",
    "01_DECISION_GO_NO_GO.md",
    "02_SCORECARD_READINESS.md",
    "03_EVIDENCIAS_TECNICAS.md",
    "04_EVIDENCIAS_UAT.md",
    "05_BRECHAS_CRITICAS.md",
    "06_PLAN_CIERRE_BRECHAS.md",
    "07_RIESGOS_Y_ACEPTACIONES.md",
    "08_RECOMENDACION_FINAL.md",
    "09_ANEXOS_TECNICOS.md",
]


def repo_root() -> Path:
    return Path(__file__).resolve().parents[4]


def package_dir() -> Path:
    return Path(__file__).resolve().parents[2]


def exports_dir() -> Path:
    return Path(__file__).resolve().parents[1]


def run_git(args: Sequence[str]) -> str:
    try:
        result = subprocess.run(
            ["git", *args],
            cwd=repo_root(),
            check=True,
            capture_output=True,
            text=True,
        )
        return result.stdout.strip()
    except Exception:
        return "No disponible"


def ensure_dirs() -> Tuple[Path, Path]:
    pdf_dir = exports_dir() / "PDFs"
    excel_dir = exports_dir() / "Excel"
    pdf_dir.mkdir(parents=True, exist_ok=True)
    excel_dir.mkdir(parents=True, exist_ok=True)
    return pdf_dir, excel_dir


def generation_date() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M")


def read_markdown(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def split_table_row(line: str) -> List[str]:
    value = line.strip().strip("|")
    return [part.strip() for part in value.split("|")]


def is_separator_row(line: str) -> bool:
    cells = split_table_row(line)
    return bool(cells) and all(re.fullmatch(r":?-{3,}:?", cell.replace(" ", "")) for cell in cells)


def markdown_blocks(markdown: str) -> Iterable[Tuple[str, object]]:
    lines = markdown.splitlines()
    i = 0
    paragraph: List[str] = []

    def flush_paragraph():
        nonlocal paragraph
        if paragraph:
            text = " ".join(part.strip() for part in paragraph if part.strip())
            paragraph = []
            if text:
                return ("paragraph", text)
        return None

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        if not stripped:
            block = flush_paragraph()
            if block:
                yield block
            i += 1
            continue
        if stripped.startswith("|") and "|" in stripped[1:]:
            block = flush_paragraph()
            if block:
                yield block
            table: List[List[str]] = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                if not is_separator_row(lines[i]):
                    table.append(split_table_row(lines[i]))
                i += 1
            if table:
                yield ("table", table)
            continue
        if stripped.startswith("#"):
            block = flush_paragraph()
            if block:
                yield block
            level = len(stripped) - len(stripped.lstrip("#"))
            text = stripped[level:].strip()
            yield (f"h{min(level, 3)}", text)
            i += 1
            continue
        if stripped.startswith("- "):
            block = flush_paragraph()
            if block:
                yield block
            items: List[str] = []
            while i < len(lines) and lines[i].strip().startswith("- "):
                items.append(lines[i].strip()[2:].strip())
                i += 1
            yield ("list", items)
            continue
        paragraph.append(stripped)
        i += 1

    block = flush_paragraph()
    if block:
        yield block


def generate_pdfs() -> None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import LETTER
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            ListFlowable,
            ListItem,
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError as exc:
        print(f"[WARN] PDF no generado: falta dependencia reportlab ({exc}).")
        return

    pdf_dir, _ = ensure_dirs()
    styles = getSampleStyleSheet()
    styles["Title"].fontName = "Helvetica-Bold"
    styles["Heading1"].spaceAfter = 8
    styles["Heading2"].spaceAfter = 6
    styles["BodyText"].fontSize = 9
    styles["BodyText"].leading = 12

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.drawString(0.55 * inch, 0.38 * inch, f"{PROJECT} - {GLOBAL_STATE}")
        canvas.drawRightString(7.95 * inch, 0.38 * inch, f"Pagina {doc.page}")
        canvas.restoreState()

    def story_for_doc(md_path: Path, include_title_page: bool = True) -> List[object]:
        markdown = read_markdown(md_path)
        story: List[object] = []
        if include_title_page:
            story.append(Paragraph(PROJECT, styles["Title"]))
            story.append(Paragraph(f"Documento: {md_path.name}", styles["Heading2"]))
            story.append(Paragraph(f"Fecha de generacion: {generation_date()}", styles["BodyText"]))
            story.append(Paragraph(f"Estado: {GLOBAL_STATE}", styles["BodyText"]))
            story.append(Spacer(1, 12))
        for kind, content in markdown_blocks(markdown):
            if kind == "h1":
                story.append(Paragraph(escape(str(content)), styles["Heading1"]))
            elif kind == "h2":
                story.append(Paragraph(escape(str(content)), styles["Heading2"]))
            elif kind == "h3":
                story.append(Paragraph(escape(str(content)), styles["Heading3"]))
            elif kind == "paragraph":
                story.append(Paragraph(escape(str(content)), styles["BodyText"]))
                story.append(Spacer(1, 4))
            elif kind == "list":
                items = [ListItem(Paragraph(escape(str(item)), styles["BodyText"])) for item in content]  # type: ignore[arg-type]
                story.append(ListFlowable(items, bulletType="bullet", leftIndent=16))
                story.append(Spacer(1, 4))
            elif kind == "table":
                raw_rows = content  # type: ignore[assignment]
                table_data = [[Paragraph(escape(str(cell)), styles["BodyText"]) for cell in row] for row in raw_rows]
                col_count = max(len(row) for row in table_data)
                widths = [7.1 * inch / col_count] * col_count
                table = Table(table_data, colWidths=widths, repeatRows=1)
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#9CA3AF")),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 7),
                            ("LEADING", (0, 0), (-1, -1), 9),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
                        ]
                    )
                )
                story.append(table)
                story.append(Spacer(1, 8))
        return story

    source_paths = [package_dir() / name for name in SOURCE_DOCS]
    for md_path in source_paths:
        if not md_path.exists():
            print(f"[WARN] Markdown no encontrado para PDF: {md_path}")
            continue
        pdf_path = pdf_dir / f"{md_path.stem}.pdf"
        doc = SimpleDocTemplate(
            str(pdf_path),
            pagesize=LETTER,
            rightMargin=0.45 * inch,
            leftMargin=0.45 * inch,
            topMargin=0.55 * inch,
            bottomMargin=0.55 * inch,
        )
        doc.build(story_for_doc(md_path), onFirstPage=footer, onLaterPages=footer)
        print(f"[OK] PDF individual: {pdf_path}")

    complete_story: List[object] = []
    for index, md_path in enumerate(source_paths):
        if not md_path.exists():
            continue
        if index > 0:
            complete_story.append(PageBreak())
        complete_story.extend(story_for_doc(md_path, include_title_page=True))
    complete_path = pdf_dir / "PAQUETE_COMITE_GO_NO_GO_COMPLETO.pdf"
    doc = SimpleDocTemplate(
        str(complete_path),
        pagesize=LETTER,
        rightMargin=0.45 * inch,
        leftMargin=0.45 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.55 * inch,
    )
    doc.build(complete_story, onFirstPage=footer, onLaterPages=footer)
    print(f"[OK] PDF completo: {complete_path}")


def generate_excel() -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError(f"No se puede generar Excel: falta openpyxl ({exc}).") from exc

    _, excel_dir = ensure_dirs()
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    def add_sheet(name: str, headers: Sequence[str], rows: Sequence[Sequence[object]]):
        ws = wb.create_sheet(title=name)
        ws.append(list(headers))
        for row in rows:
            ws.append(list(row))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
                cell.border = border
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)
        return ws

    branch = run_git(["rev-parse", "--abbrev-ref", "HEAD"])
    commit = run_git(["rev-parse", "--short", "HEAD"])

    resumen_headers = [
        "Proyecto",
        "Versión paquete",
        "Fecha generación",
        "Rama",
        "Commit",
        "Estado global",
        "Recomendación",
        "Scorecard",
        "Responsable ejecución",
        "Ambiente",
        "Observaciones",
    ]
    add_sheet(
        "Resumen",
        resumen_headers,
        [
            [
                PROJECT,
                PACKAGE_VERSION,
                generation_date(),
                branch,
                commit,
                GLOBAL_STATE,
                RECOMMENDATION,
                SCORECARD,
                "Pendiente asignar",
                ENVIRONMENT,
                "Productivo: NO-GO. No contiene secretos ni datos reales.",
            ]
        ],
    )

    checklist_headers = [
        "ID",
        "Categoría",
        "Control",
        "Descripción",
        "Responsable",
        "Perfil requerido",
        "Estado",
        "Evidencia requerida",
        "Evidencia obtenida",
        "Defecto asociado",
        "Observación",
    ]
    checklist_rows = [
        ("CHK-001", "Acceso y autenticacion", "Usuario puede ingresar a la SPA", "Validar disponibilidad de la SPA Docker.", "QA/UAT", "Admin u operador", "Pendiente", "Captura pantalla", "", "", ""),
        ("CHK-002", "Acceso y autenticacion", "Login con usuario aprobado", "Login con usuario demo aprobado sin documentar password.", "QA/UAT", "Admin", "Pendiente", "Request/response enmascarado", "", "", ""),
        ("CHK-003", "Navegacion", "Menu carga correctamente", "Validar menu autenticado.", "QA/UAT", "Admin", "Pendiente", "Captura pantalla", "", "", ""),
        ("CHK-004", "Roles y permisos", "Usuario visualiza opciones segun rol", "Comparar menu visible contra rol esperado.", "Seguridad/QA", "Admin", "OK", "Captura pantalla/login sanitizado", "Login/JWT evidencia Admin y ACH.Operator", "DEF-UAT-015", "Cerrado para UAT controlado."),
        ("CHK-005", "Roles y permisos", "Usuario no accede a opciones no permitidas", "Validar rechazo controlado.", "Seguridad/QA", "Admin/Operador", "Pendiente", "Request/response", "", "", ""),
        ("CHK-006", "Navegacion", "Dashboard carga", "Validar pantalla inicial o home.", "QA/UAT", "Admin", "Pendiente", "Captura pantalla", "", "", ""),
        ("CHK-007", "Transacciones", "Transacciones carga", "Validar modulo de transacciones.", "QA/UAT", "Admin u operador", "Pendiente", "Captura pantalla", "", "", ""),
        ("CHK-008", "Datos maestros", "Datos maestros carga", "Validar modulo general de datos maestros.", "QA/UAT", "Admin", "Pendiente", "Captura pantalla", "", "", ""),
        ("CHK-009", "Datos maestros", "Clearing houses carga", "Validar camaras de compensacion sinteticas.", "QA/UAT", "Admin", "Pendiente", "Captura pantalla/API", "", "", ""),
        ("CHK-010", "Datos maestros", "Financial institutions carga", "Validar instituciones financieras sinteticas.", "QA/UAT", "Admin", "Pendiente", "Captura pantalla/API", "", "", ""),
        ("CHK-011", "Datos maestros", "ACH cycles carga", "Validar ciclos ACH.", "QA/UAT", "Admin", "Pendiente", "Captura pantalla/API", "", "", ""),
        ("CHK-012", "Datos maestros", "Company entry descriptions carga", "Validar descripciones de entrada.", "QA/UAT", "Admin", "Pendiente", "Captura pantalla/API", "", "", ""),
        ("CHK-013", "Transacciones", "Consulta de transacciones funciona", "Consultar transacciones sinteticas.", "QA/UAT", "Admin u operador", "Pendiente", "Captura pantalla/API", "", "", ""),
        ("CHK-014", "Transacciones", "Creacion de transaccion sintetica controlada", "Crear una transaccion con datos sinteticos.", "QA/UAT", "Admin u operador", "Pendiente", "Request/response", "", "", ""),
        ("CHK-015", "Transacciones", "Validacion de duplicado/idempotencia", "Reintentar payload sintetico duplicado.", "QA/UAT", "Admin u operador", "Pendiente", "Request/response", "", "DEF-UAT-018", "Contrato actual documentado."),
        ("CHK-016", "Auditoria", "Trazabilidad de transaccion", "Consultar historial o evidencia de auditoria.", "QA/Auditoria", "Admin", "Pendiente", "Registro API/BD", "", "DEF-UAT-017", "Cerrado para nuevas transacciones."),
        ("CHK-017", "Auditoria", "Evento inicial de transaccion", "Validar evento Pending -> Pending CREATED.", "QA/Auditoria", "Admin", "Pendiente", "Registro API/BD", "", "DEF-UAT-017", ""),
        ("CHK-018", "Reportes", "Reporte basico", "Validar reporte disponible sin datos reales.", "QA/UAT", "Admin", "Pendiente", "Captura pantalla", "", "", ""),
        ("CHK-019", "Operacion", "Logs sin errores criticos", "Revisar logs API/SPA/PostgreSQL.", "SRE/QA", "Tecnico", "Pendiente", "Log API/SPA", "", "", ""),
        ("CHK-020", "Evidencias", "Evidencias adjuntas", "Adjuntar evidencia por prueba.", "QA/UAT", "QA", "Pendiente", "Ruta/hash", "", "", ""),
        ("CHK-021", "Evidencias", "Defectos registrados", "Registrar defectos de UAT.", "QA/UAT", "QA", "Pendiente", "Matriz defectos", "", "", ""),
        ("CHK-022", "Actas", "Acta preliminar generada", "Preparar acta de resultados UAT.", "QA/PMO", "QA/PMO", "Pendiente", "Acta", "", "", ""),
    ]
    ws_checklist = add_sheet("Checklist_UAT_Operativo", checklist_headers, checklist_rows)

    pruebas_headers = [
        "ID Prueba",
        "Modulo",
        "Escenario",
        "Objetivo",
        "Perfil usuario",
        "Precondiciones",
        "Datos de prueba",
        "Pasos",
        "Resultado esperado",
        "Resultado obtenido",
        "Estado",
        "Evidencia",
        "Defecto",
        "Prioridad",
        "Observaciones",
    ]
    prueba_specs = [
        ("UAT-OP-001", "Acceso", "Login operador/admin", "Validar ingreso al sistema.", "Admin", "SPA disponible", "Usuario demo admin sin password documentado", "Abrir SPA; ingresar usuario; autenticar.", "Login exitoso o error controlado.", "", "Pendiente", "", "", "Alta", ""),
        ("UAT-OP-002", "Navegacion", "Carga menu", "Validar menu autenticado.", "Admin", "Login exitoso", "Token en memoria no documentado", "Ingresar y esperar menu.", "Menu visible.", "", "Pendiente", "", "", "Alta", ""),
        ("UAT-OP-003", "Dashboard", "Consulta dashboard/home", "Validar pantalla inicial.", "Admin", "Login exitoso", "N/A", "Navegar a home.", "Pantalla carga sin 500.", "", "Pendiente", "", "", "Media", ""),
        ("UAT-OP-004", "Datos maestros", "Consulta instituciones financieras", "Validar endpoint/pantalla.", "Admin", "Login exitoso", "Bancos sinteticos", "Abrir modulo.", "Respuesta API valida.", "", "Pendiente", "", "", "Alta", ""),
        ("UAT-OP-005", "Datos maestros", "Consulta camaras de compensacion", "Validar clearing houses.", "Admin", "Login exitoso", "Camaras sinteticas", "Abrir modulo.", "Respuesta API valida.", "", "Pendiente", "", "", "Alta", ""),
        ("UAT-OP-006", "Datos maestros", "Consulta ciclos ACH", "Validar ciclos.", "Admin", "Login exitoso", "Ciclo sintetico", "Abrir modulo.", "Respuesta API valida.", "", "Pendiente", "", "", "Alta", ""),
        ("UAT-OP-007", "Datos maestros", "Consulta company entry descriptions", "Validar catalogo.", "Admin", "Login exitoso", "Descripcion sintetica", "Abrir modulo.", "Respuesta API valida.", "", "Pendiente", "", "", "Media", ""),
        ("UAT-OP-008", "Transacciones", "Consulta transacciones", "Validar listado.", "Admin/Operador", "Login exitoso", "Transacciones sinteticas", "Abrir listado.", "Listado carga sin datos reales.", "", "Pendiente", "", "", "Alta", ""),
        ("UAT-OP-009", "Transacciones", "Crear transaccion sintetica", "Crear transaccion controlada.", "Admin/Operador", "Datos maestros disponibles", "Referencia UAT-SINT", "Crear payload sintetico.", "Creacion exitosa sin 500.", "", "Pendiente", "", "", "Alta", ""),
        ("UAT-OP-010", "Transacciones", "Validar duplicado transaccion sintetica", "Verificar deduplicacion.", "Admin/Operador", "Transaccion creada", "Misma referencia/payload", "Reintentar payload.", "Error JSON controlado; no duplica.", "", "Pendiente", "", "DEF-UAT-018", "Alta", ""),
        ("UAT-OP-011", "Auditoria", "Consultar trazabilidad", "Validar historial.", "Admin", "Transaccion creada", "TransactionId sintetico", "Consultar trazabilidad.", "Eventos disponibles.", "", "Pendiente", "", "DEF-UAT-017", "Alta", ""),
        ("UAT-OP-012", "Auditoria", "Validar evento inicial", "Validar Pending -> Pending CREATED.", "Admin", "Transaccion nueva", "TransactionId sintetico", "Consultar evento.", "Evento inicial presente.", "", "Pendiente", "", "DEF-UAT-017", "Alta", ""),
        ("UAT-OP-013", "ACH", "Consultar respuestas ACH", "Validar respuestas/catalogos.", "Admin", "Login exitoso", "Catalogo sintetico", "Abrir modulo.", "Respuesta API valida.", "", "Pendiente", "", "", "Media", ""),
        ("UAT-OP-014", "Reportes", "Consultar reportes", "Validar reporte basico.", "Admin", "Login exitoso", "Datos sinteticos", "Abrir reporte.", "Reporte carga.", "", "Pendiente", "", "", "Media", ""),
        ("UAT-OP-015", "Errores", "Validar error controlado", "Forzar caso invalido seguro.", "Admin", "Login exitoso", "Payload sintetico invalido", "Enviar peticion invalida.", "400/401/403 controlado, no 500.", "", "Pendiente", "", "", "Alta", ""),
        ("UAT-OP-016", "Acceso", "Validar logout", "Cerrar sesion.", "Admin", "Sesion activa", "N/A", "Ejecutar logout.", "Sesion cerrada.", "", "Pendiente", "", "", "Media", ""),
        ("UAT-OP-017", "Roles", "Validar rol Admin", "Confirmar permisos Admin.", "Admin", "Login exitoso", "Usuario admin", "Revisar menu y endpoints.", "Permisos Admin visibles.", "", "Pendiente", "", "", "Alta", ""),
        ("UAT-OP-018", "Roles", "Validar rol ACH.Operator", "Validar rol operador con usuario demo multirol.", "Admin + ACH.Operator", "Seed/migracion aplicada", "Usuario demo admin sin password documentado", "Login y revisar claims/menu.", "Rol visible y autorizado.", "Admin y ACH.Operator visibles en respuesta/JWT sanitizados.", "Ejecutada OK", "EV-FUNC-039", "DEF-UAT-015", "Alta", "Cerrado para UAT controlado; evaluar usuario operador separado antes de productivo."),
        ("UAT-OP-019", "NACHA-M", "Validar NACHA-M layouts", "Revisar layout general.", "Admin/QA", "Endpoint/proxy OK", "Archivo sintetico", "Ejecutar validacion.", "Resultado conforme.", "", "Pendiente", "", "DEF-UAT-020", "Alta", ""),
        ("UAT-OP-020", "NACHA-M", "Validar NACHA-M registro 1 sintetico", "Validar header.", "Admin/QA", "Archivo sintetico", "Registro 1", "Validar campos.", "Conforme o defecto.", "", "Pendiente", "", "DEF-UAT-020", "Alta", ""),
        ("UAT-OP-021", "NACHA-M", "Validar NACHA-M registro 5 sintetico", "Validar batch header.", "Admin/QA", "Archivo sintetico", "Registro 5", "Validar campos.", "Conforme o defecto.", "", "Pendiente", "", "DEF-UAT-020", "Alta", ""),
        ("UAT-OP-022", "NACHA-M", "Validar NACHA-M registro 6 sintetico", "Validar entry detail.", "Admin/QA", "Archivo sintetico", "Registro 6", "Validar campos.", "Conforme o defecto.", "", "Pendiente", "", "DEF-UAT-020", "Alta", ""),
        ("UAT-OP-023", "NACHA-M", "Validar NACHA-M registro 7 sintetico", "Validar addenda.", "Admin/QA", "Archivo sintetico", "Registro 7", "Validar campos.", "Conforme o defecto.", "", "Pendiente", "", "DEF-UAT-020", "Alta", ""),
        ("UAT-OP-024", "NACHA-M", "Validar NACHA-M registro 8 sintetico", "Validar batch control.", "Admin/QA", "Archivo sintetico", "Registro 8", "Validar campos.", "Conforme o defecto.", "", "Pendiente", "", "DEF-UAT-020", "Alta", ""),
        ("UAT-OP-025", "NACHA-M", "Validar NACHA-M registro 9 sintetico", "Validar file control.", "Admin/QA", "Archivo sintetico", "Registro 9", "Validar campos.", "Conforme o defecto.", "", "Pendiente", "", "DEF-UAT-020", "Alta", ""),
        ("UAT-OP-026", "Conciliacion", "Validar conciliacion basica", "Validar conteos/montos sinteticos.", "Admin/QA", "Datos sinteticos", "Transacciones sinteticas", "Ejecutar consulta.", "Totales coherentes.", "", "Pendiente", "", "", "Media", ""),
        ("UAT-OP-027", "CENIT", "Validar CENIT/CUD", "Validar integracion o waiver.", "Admin/QA", "Alcance definido", "Datos sinteticos", "Ejecutar segun alcance.", "Evidencia o pendiente formal.", "", "Bloqueada", "", "CENIT-CUD", "Alta", "Pendiente si no aplica aun."),
        ("UAT-OP-028", "Seguridad", "Validar sobre digital/certificados", "Validar flujo sin certificados privados.", "Seguridad", "Alcance definido", "Certificados de prueba", "Ejecutar prueba segura.", "Evidencia o waiver.", "", "Bloqueada", "", "SOBRE-DIGITAL", "Alta", "No usar certificados productivos."),
        ("UAT-OP-029", "Operacion", "Validar backup/restore", "Ejecutar recuperacion controlada.", "SRE", "Ambiente controlado", "Backup de prueba", "Ejecutar restore/rollback.", "Recuperacion documentada.", "", "Bloqueada", "", "BKP-RESTORE", "Alta", "Pendiente."),
        ("UAT-OP-030", "Actas", "Validar acta y cierre de evidencias", "Cerrar paquete UAT.", "QA/PMO", "Evidencias completas", "N/A", "Revisar acta.", "Acta firmada o pendiente formal.", "", "Pendiente", "", "ACTAS", "Alta", ""),
    ]
    ws_pruebas = add_sheet("Set_Pruebas_Operativas", pruebas_headers, prueba_specs)

    evidencia_headers = ["ID Evidencia", "ID Prueba", "Tipo Evidencia", "Descripción", "Ruta/Referencia", "Hash SHA256", "Responsable", "Fecha", "Estado", "Observación"]
    evidencia_rows = [
        ("EVI-001", "UAT-OP-001", "Captura pantalla", "Login SPA", "", "", "QA/UAT", "", "Pendiente", "No incluir password."),
        ("EVI-002", "UAT-OP-010", "Request/response", "Duplicado controlado", "", "", "QA/UAT", "", "Pendiente", "No incluir token completo."),
        ("EVI-003", "UAT-OP-012", "Registro BD", "Evento inicial", "", "", "QA/Auditoria", "", "Pendiente", "Solo datos sinteticos."),
        ("EVI-004", "UAT-OP-030", "Acta", "Acta UAT", "", "", "PMO/QA", "", "Pendiente", ""),
    ]
    ws_evid = add_sheet("Evidencias", evidencia_headers, evidencia_rows)

    defectos_headers = ["ID Defecto", "ID Prueba", "Módulo", "Severidad", "Descripción", "Evidencia", "Responsable", "Estado", "Fecha apertura", "Fecha cierre", "Decisión", "Observación"]
    defectos_rows = [
        ("DEF-UAT-015", "UAT-OP-018", "Roles", "Alta", "ACH.Operator no asignado/no visible.", "EV-FUNC-039", "Seguridad/Backend", "Cerrado", "2026-05-18", "2026-05-19", "Opcion A: admin demo multirol para UAT controlado", "No debilita auth; evaluar operador separado para productivo."),
        ("DEF-UAT-020", "UAT-OP-019", "NACHA-M", "Bloqueante", "NACHA-M 1/5/6/7/8/9 pendiente campo-a-campo.", "", "Arquitectura ACH/QA", "Abierto", "", "", "Pendiente", ""),
        ("CENIT-CUD", "UAT-OP-027", "CENIT", "Bloqueante", "CENIT/CUD pendiente.", "", "Integracion/Negocio", "Abierto", "", "", "Pendiente", ""),
        ("SOBRE-DIGITAL", "UAT-OP-028", "Seguridad", "Bloqueante", "Sobre digital/firma/certificados pendiente.", "", "Seguridad/Integracion", "Abierto", "", "", "Pendiente", ""),
        ("BKP-RESTORE", "UAT-OP-029", "Operacion", "Bloqueante", "Backup/restore/rollback pendiente.", "", "Operaciones/SRE", "Abierto", "", "", "Pendiente", ""),
    ]
    ws_def = add_sheet("Defectos", defectos_headers, defectos_rows)

    datos_headers = ["ID Dato", "Tipo", "Valor sintético", "Uso", "Sensibilidad", "Permitido en Git", "Observación"]
    datos_rows = [
        ("DAT-001", "Usuario demo", "admin", "Login UAT controlado", "Baja sin password", "Si", "No incluir password."),
        ("DAT-002", "Cliente", "Cliente UAT Sintetico", "Transaccion sintetica", "Baja", "Si", "No representa cliente real."),
        ("DAT-003", "Documento", "999999999", "Identificacion sintetica", "Baja", "Si", "No usar identificaciones reales."),
        ("DAT-004", "Cuenta origen", "0000000001", "Transaccion sintetica", "Baja", "Si", "No usar cuenta real."),
        ("DAT-005", "Cuenta destino", "0000000002", "Transaccion sintetica", "Baja", "Si", "No usar cuenta real."),
        ("DAT-006", "Banco origen", "Banco UAT Origen", "Datos maestros sinteticos", "Baja", "Si", "No banco productivo real."),
        ("DAT-007", "Banco destino", "Banco UAT Destino", "Datos maestros sinteticos", "Baja", "Si", "No banco productivo real."),
        ("DAT-008", "Referencia", "UAT-SINT", "Pruebas sinteticas", "Baja", "Si", "Prefijo controlado."),
        ("DAT-009", "Monto", "1000", "Monto pequeno sintetico", "Baja", "Si", "Sin impacto real."),
        ("DAT-010", "Archivo NACHA-M", "Archivo sintetico pendiente", "Validacion campo-a-campo", "Media", "Si", "No usar NACHA-M productivo real."),
    ]
    add_sheet("Datos_Sinteticos", datos_headers, datos_rows)

    criterios_headers = ["Criterio", "Estado actual", "Evidencia", "Bloquea productivo", "Condición para GO", "Responsable", "Observación"]
    criterios_rows = [
        ("CI backend", "OK", "dotnet-ci", "No", "Mantener OK", "Tecnologia", ""),
        ("CI Angular", "OK", "angular-ci", "No", "Mantener OK", "Tecnologia", ""),
        ("Runtime Docker", "OK", "Docker runtime", "No", "Mantener runtime estable", "DevOps/SRE", ""),
        ("UAT tecnico", "OK con observaciones", "Docs UAT tecnico", "No", "Sin regresiones", "QA", ""),
        ("UAT funcional", "PARCIALMENTE OK", "Docs UAT funcional", "Si", "UAT formal firmado", "QA/Negocio", ""),
        ("NACHA-M campo-a-campo", "PENDIENTE", "Matriz NACHA-M", "Si", "Registros 1/5/6/7/8/9 validados", "Arquitectura ACH", ""),
        ("CENIT/CUD", "PENDIENTE", "Brechas", "Si", "Validacion o waiver", "Integracion", ""),
        ("Sobre digital", "PENDIENTE", "Brechas", "Si", "Firma/certificados validados", "Seguridad", ""),
        ("Certificados", "PENDIENTE", "Revision seguridad", "Si", "Certificados de prueba y custodia aprobada", "Seguridad", ""),
        ("Seguridad", "PARCIAL", "Revision seguridad", "Si", "Brechas cerradas o aceptadas formalmente", "Seguridad", "ACH.Operator cerrado para UAT; siguen secretos/certificados/OpenBao segun alcance."),
        ("OpenBao/secrets", "PENDIENTE", "Brechas", "Si", "Gestion de secretos validada", "DevOps/Security", ""),
        ("Backup/restore", "PENDIENTE", "Brechas", "Si", "Prueba aprobada", "SRE", ""),
        ("Actas", "PENDIENTE", "Actas UAT", "Si", "Firmas completas", "PMO/QA", ""),
        ("Aprobaciones", "PENDIENTE", "Comite", "Si", "Aprobaciones formales", "Direccion", "Productivo: NO-GO."),
    ]
    ws_criterios = add_sheet("Criterios_GO_NO_GO", criterios_headers, criterios_rows)

    firmas_headers = ["Area", "Responsable", "Rol", "Decision", "Fecha", "Firma/Referencia", "Observacion"]
    firmas_rows = [
        ("Negocio", "", "Aprobador funcional", "Pendiente", "", "", ""),
        ("Operaciones", "", "Aprobador operativo", "Pendiente", "", "", ""),
        ("Tecnologia", "", "Aprobador tecnico", "Pendiente", "", "", ""),
        ("Seguridad", "", "Aprobador seguridad", "Pendiente", "", "", ""),
        ("Auditoria", "", "Revisor auditoria", "Pendiente", "", "", ""),
        ("Infraestructura", "", "Aprobador infraestructura", "Pendiente", "", "", ""),
        ("Soporte", "", "Aprobador soporte", "Pendiente", "", "", ""),
        ("Proveedor/camara si aplica", "", "Contraparte", "Pendiente", "", "", "Solo si aplica al alcance."),
    ]
    ws_firmas = add_sheet("Firmas_Aprobaciones", firmas_headers, firmas_rows)

    def add_validation(ws, col_letter: str, allowed: Sequence[str], start_row: int = 2, end_row: int = 500):
        dv = DataValidation(type="list", formula1=f'"{",".join(allowed)}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")

    checklist_states = ["Pendiente", "OK", "Falló", "Bloqueado", "No aplica"]
    test_states = ["Pendiente", "Ejecutada OK", "Falló", "Bloqueada", "No aplica"]
    evidence_states = ["Pendiente", "Adjunta", "Rechazada", "No aplica"]
    defect_states = ["Abierto", "En análisis", "Corregido", "Rechazado", "Diferido", "Aceptado como riesgo", "Cerrado"]
    severities = ["Bloqueante", "Alta", "Media", "Baja"]
    priorities = ["Alta", "Media", "Baja"]
    decisions = ["Aprueba", "Aprueba con observaciones", "Rechaza", "Pendiente"]

    add_validation(ws_checklist, "G", checklist_states)
    add_validation(ws_pruebas, "K", test_states)
    add_validation(ws_pruebas, "N", priorities)
    add_validation(ws_evid, "C", ["Captura pantalla", "Request/response", "Log API", "Log SPA", "Registro BD", "Archivo sintético", "Acta", "Aprobación", "Defecto", "Otro"])
    add_validation(ws_evid, "I", evidence_states)
    add_validation(ws_def, "D", severities)
    add_validation(ws_def, "H", defect_states)
    add_validation(ws_firmas, "D", decisions)

    status_fills = {
        "OK": "C6EFCE",
        "Ejecutada OK": "C6EFCE",
        "Pendiente": "FFF2CC",
        "Bloqueado": "F8CBAD",
        "Bloqueada": "F8CBAD",
        "Falló": "FFC7CE",
        "No aplica": "D9EAD3",
        "Abierto": "FFC7CE",
        "Cerrado": "C6EFCE",
        "Corregido": "C6EFCE",
    }
    for ws, col in [(ws_checklist, "G"), (ws_pruebas, "K"), (ws_evid, "I"), (ws_def, "H")]:
        for value, color in status_fills.items():
            ws.conditional_formatting.add(
                f"{col}2:{col}500",
                CellIsRule(operator="equal", formula=[f'"{value}"'], fill=PatternFill("solid", fgColor=color)),
            )

    output = excel_dir / "CHECKLIST_UAT_OPERATIVO_ACH_INTERBANK.xlsx"
    wb.save(output)
    print(f"[OK] Excel operativo: {output}")


def main() -> int:
    print(f"[INFO] Generando exportables para {PROJECT}")
    print(f"[INFO] Estado: {GLOBAL_STATE}")
    ensure_dirs()
    generate_excel()
    generate_pdfs()
    print("[OK] Generacion finalizada.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        raise SystemExit(1)
