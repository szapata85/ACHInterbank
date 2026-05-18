#!/usr/bin/env python3
from pathlib import Path
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
EXPORTS = ROOT / "docs/uat/operator-guides/exports"
PDF_PATH = EXPORTS / "UAT_ACHInterbank_Guia_Operativa_Usuarios.pdf"
XLSX_PATH = EXPORTS / "UAT_ACHInterbank_Set_Pruebas_Operativas.xlsx"


def build_pdf(path: Path) -> None:
    c = canvas.Canvas(str(path), pagesize=A4)
    width, height = A4
    sections = [
        ("UAT ACHInterbank — Guía Operativa para Usuarios No Técnicos", [
            "Documento derivado del paquete 12B y protocolo 12A.",
            "Uso: ejecución UAT funcional con datos reales o anonimizados.",
        ]),
        ("Objetivo", ["Guiar a usuarios no técnicos para ejecutar casos UAT, registrar evidencias, reportar defectos y preparar aprobación humana."]),
        ("Alcance", ["Cobertura S1-10, S1-11, S1-12, S1-13 y S1-20.", "Aplicable a operaciones, tesorería, compliance, riesgo y QA UAT funcional."]),
        ("Guía de ejecución", [
            "1) Recibir caso UAT.", "2) Confirmar datos autorizados.", "3) Ejecutar operación o consultar resultado.",
            "4) Comparar esperado vs obtenido.", "5) Guardar evidencia.", "6) Registrar defecto si aplica.",
            "7) Marcar estado del caso.", "8) Solicitar aprobación.",
        ]),
        ("Estados permitidos", ["Pendiente, En ejecución, Aprobado, Aprobado con observaciones, Rechazado, Bloqueado."]),
        ("Protección de datos sensibles", [
            "No incluir datos reales sensibles, cuentas, identificaciones completas ni saldos completos.",
            "No incluir PFX, passwords, llaves privadas ni certificados privados.",
            "Usar enmascaramiento, hash y referencia interna; custodiar soportes en ubicación segura aprobada.",
        ]),
        ("Casos UAT resumidos", [
            "S1-10: neteo por ciclo, posiciones por participante y reproceso sin duplicidad.",
            "S1-11: separación liquidez simulada vs saldo real CUD, evidencia y conciliación CUD.",
            "S1-12: naming ACH/CENIT/devoluciones-ROR y validación de cámara correcta.",
            "S1-13: validación de firma/cifrado, recepción externa y rechazo controlado.",
            "S1-20: runbook, checklist, acta UAT y defectos cerrados/aceptados.",
        ]),
        ("Evidencia aceptada / no aceptada", [
            "Aceptada: captura enmascarada, PDF reporte, CSV/Excel exportado, hash/referencia, acta, aprobación trazable.",
            "No aceptada: captura sin contexto, aprobación verbal, evidencia con datos sensibles visibles, material criptográfico privado.",
        ]),
        ("Reporte de defectos", ["Severidades permitidas: P0, P1, P2, P3.", "P0 abierto bloquea GO UAT formal.", "P1 requiere workaround aprobado."]),
        ("Aprobación y firma", ["Aprobadores mínimos: QA UAT, Operaciones, Tesorería, Seguridad, Compliance, Riesgo Operacional, Dueño de proceso, Tecnología."]),
        ("Advertencia", ["Este documento no habilita producción. GO productivo: NO. NO-GO productivo vigente hasta scorecard y aprobación formal."]),
    ]

    y = height - 2 * cm
    c.setFont("Helvetica-Bold", 16)
    c.drawString(2 * cm, y, sections[0][0])
    y -= 1 * cm
    c.setFont("Helvetica", 10)
    for line in sections[0][1]:
        c.drawString(2 * cm, y, line)
        y -= 0.55 * cm

    for title, lines in sections[1:]:
        if y < 3 * cm:
            c.showPage()
            y = height - 2 * cm
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2 * cm, y, title)
        y -= 0.6 * cm
        c.setFont("Helvetica", 10)
        for line in lines:
            if y < 2.5 * cm:
                c.showPage()
                y = height - 2 * cm
                c.setFont("Helvetica", 10)
            c.drawString(2.4 * cm, y, f"- {line}")
            y -= 0.5 * cm
        y -= 0.2 * cm

    c.save()


def setup_sheet(ws, headers, widths):
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(vertical="top", wrap_text=True)
    ws.append(headers)
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        ws.column_dimensions[get_column_letter(idx)].width = widths[idx - 1]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_xlsx(path: Path) -> None:
    wb = Workbook()
    center = Alignment(vertical="top", wrap_text=True)

    ws = wb.active
    ws.title = "Instrucciones"
    setup_sheet(ws, ["Campo", "Descripción", "Responsable"], [26, 80, 26])
    ws.append(["Uso del archivo", "Plantilla operativa para ejecución UAT sin datos sensibles.", "QA UAT / Operaciones"])
    ws.append(["Regla de seguridad", "No incluir datos reales sensibles, PFX, llaves privadas ni passwords.", "Todos"])
    ws.append(["Estado inicial", "Todos los casos deben iniciar en Pendiente.", "QA UAT"])
    ws.append(["GO productivo", "Debe mantenerse en NO hasta aprobación formal y scorecard.", "Comité UAT"])

    casos = wb.create_sheet("Casos_UAT")
    case_headers = ["ID Caso", "Dominio S1", "Cámara", "Nombre del caso", "Objetivo", "Datos requeridos", "Pasos operativos", "Resultado esperado", "Resultado obtenido", "Estado", "Evidencia requerida", "ID evidencia", "Defecto asociado", "Aprobador", "Observaciones"]
    setup_sheet(casos, case_headers, [20, 12, 14, 30, 35, 28, 42, 35, 30, 16, 28, 16, 18, 24, 24])
    initial_cases = [
        ("UAT-OP-S1-10-001", "S1-10", "CENIT", "Validar neteo por ciclo CENIT", "Confirmar totales consistentes y trazables.", "Ciclo/reporte/control autorizado", "Revisar ciclo, totales y comparación de control.", "Neteo consistente por ciclo/participante/posición.", "", "Pendiente", "Reporte + captura + referencia", "EV-001", "", "Operaciones + Tesorería", ""),
        ("UAT-OP-S1-11-001", "S1-11", "CENIT/CUD", "Separar liquidez simulada vs CUD real", "Evitar tratar liquidez simulada como saldo real CUD.", "Reporte liquidez + soporte CUD", "Comparar y registrar separación explícita.", "Liquidez simulada no equivale a saldo real CUD.", "", "Pendiente", "Comparativo + referencia CUD", "EV-002", "", "Tesorería + Riesgo", ""),
        ("UAT-OP-S1-12-001", "S1-12", "ACH", "Validar naming ACH", "Confirmar nombre de archivo ACH correcto.", "Regla de naming + archivo", "Comparar esperado vs obtenido.", "Nombre coincide con la regla.", "", "Pendiente", "Captura de nombre + regla", "EV-003", "", "Operaciones ACH + Compliance", ""),
        ("UAT-OP-S1-13-001", "S1-13", "ACH/CENIT", "Validar firmado/cifrado saliente", "Validar archivo saliente con control operativo.", "Archivo y validación operativa", "Revisar validación y registrar evidencia.", "Validación exitosa de firma/cifrado.", "", "Pendiente", "Constancia de validación", "EV-004", "", "Seguridad + Operaciones", ""),
        ("UAT-OP-S1-20-001", "S1-20", "Ambas", "Ejecutar runbook operativo", "Confirmar ejecución completa del runbook.", "Runbook vigente + caso", "Ejecutar pasos y registrar hitos.", "Runbook ejecutado y documentado.", "", "Pendiente", "Bitácora + checklist", "EV-005", "", "Operaciones + QA UAT", ""),
    ]
    for row in initial_cases:
        casos.append(row)

    evid = wb.create_sheet("Evidencias")
    evid_headers = ["ID Evidencia", "ID Caso", "Dominio S1", "Cámara", "Tipo de evidencia", "Descripción", "Hash / referencia", "Ubicación segura", "¿Datos enmascarados?", "Responsable", "Fecha", "Estado", "Observaciones"]
    setup_sheet(evid, evid_headers, [16, 20, 12, 14, 20, 30, 24, 28, 20, 22, 14, 16, 24])
    evid.append(["EV-001", "UAT-OP-S1-10-001", "S1-10", "CENIT", "PDF de reporte", "Reporte de neteo enmascarado", "REF-NETEO-001", "Repositorio documental seguro", "Sí", "Operaciones", "", "Pendiente", ""])

    defs = wb.create_sheet("Defectos")
    def_headers = ["ID Defecto", "ID Caso", "Dominio S1", "Cámara", "Descripción", "Resultado esperado", "Resultado obtenido", "Severidad", "Impacto operativo", "¿Bloquea aprobación?", "Responsable", "Fecha objetivo", "Estado", "Workaround", "Observaciones"]
    setup_sheet(defs, def_headers, [14, 20, 12, 14, 28, 28, 28, 12, 22, 20, 20, 16, 16, 20, 24])
    defs.append(["", "", "", "", "", "", "", "P0", "", "Sí", "", "", "Pendiente", "", ""])

    appr = wb.create_sheet("Aprobadores")
    app_headers = ["Rol", "Nombre", "Área", "Dominio que aprueba", "Decisión", "Fecha", "Firma / trazabilidad", "Observaciones"]
    setup_sheet(appr, app_headers, [20, 22, 22, 22, 22, 14, 26, 24])
    for role, dom in [("QA UAT", "Todos"), ("Operaciones", "Todos"), ("Tesorería", "S1-10/S1-11"), ("Seguridad", "S1-13"), ("Compliance", "Todos"), ("Riesgo Operacional", "Todos"), ("Dueño de proceso", "Cierre funcional"), ("Tecnología", "Soporte")]:
        appr.append([role, "", "", dom, "Pendiente", "", "", ""])

    summary = wb.create_sheet("Resumen_Ejecucion")
    setup_sheet(summary, ["Métrica", "Valor"], [40, 30])
    for metric, value in [("Total casos", "5"), ("Pendientes", "5"), ("En ejecución", "0"), ("Aprobados", "0"), ("Aprobados con observaciones", "0"), ("Rechazados", "0"), ("Bloqueados", "0"), ("P0 abiertos", "0"), ("P1 abiertos", "0"), ("P2/P3 abiertos", "0"), ("GO UAT formal", "Pendiente"), ("GO productivo", "NO"), ("NO-GO productivo", "Vigente")]:
        summary.append([metric, value])

    score = wb.create_sheet("Scorecard_UAT")
    score_headers = ["Dominio", "Bloqueante", "Estado", "Evidencia mínima", "Aprobación requerida", "Decisión", "Observaciones"]
    setup_sheet(score, score_headers, [14, 30, 16, 34, 34, 20, 28])
    for row in [
        ("S1-10", "Neteo CENIT E2E", "Pendiente", "Reporte neteo + trazabilidad", "Operaciones + Tesorería", "Pendiente", ""),
        ("S1-11", "Liquidez/CUD", "Pendiente", "Soporte CUD + conciliación", "Tesorería + Riesgo", "Pendiente", ""),
        ("S1-12", "Naming externo", "Pendiente", "Nombre esperado vs obtenido", "Operaciones + Compliance", "Pendiente", ""),
        ("S1-13", "Sobre/firma/cifrado", "Pendiente", "Validación de firma/cifrado", "Seguridad + Operaciones", "Pendiente", ""),
        ("S1-20", "UAT/runbooks/evidencia", "Pendiente", "Checklist + acta + defectos", "Comité UAT", "Pendiente", ""),
    ]:
        score.append(row)

    states = "Pendiente,En ejecución,Aprobado,Aprobado con observaciones,Rechazado,Bloqueado"
    decision = "Pendiente,Aprobado,Aprobado con observaciones,Rechazado"
    severities = "P0,P1,P2,P3"

    dv_state = DataValidation(type="list", formula1=f'"{states}"', allow_blank=True)
    casos.add_data_validation(dv_state); dv_state.add("J2:J500")
    defs.add_data_validation(dv_state); dv_state.add("M2:M500")
    score.add_data_validation(dv_state); dv_state.add("C2:C500")

    dv_sev = DataValidation(type="list", formula1=f'"{severities}"', allow_blank=True)
    defs.add_data_validation(dv_sev); dv_sev.add("H2:H500")

    dv_yes_no = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
    defs.add_data_validation(dv_yes_no); dv_yes_no.add("J2:J500")

    dv_dec = DataValidation(type="list", formula1=f'"{decision}"', allow_blank=True)
    appr.add_data_validation(dv_dec); dv_dec.add("E2:E500")
    score.add_data_validation(dv_dec); dv_dec.add("F2:F500")

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = center

    wb.save(path)


def main() -> None:
    EXPORTS.mkdir(parents=True, exist_ok=True)
    build_pdf(PDF_PATH)
    build_xlsx(XLSX_PATH)
    print(f"Generated: {PDF_PATH}")
    print(f"Generated: {XLSX_PATH}")


if __name__ == "__main__":
    main()
